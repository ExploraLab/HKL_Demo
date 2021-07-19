# Common Functions Libraries
# --------------------------------------------------------
import json
import os
import datetime

import openpyxl

from OCR_Config import ConfigParam

config = ConfigParam()
rootFolder = config.localRootFolder
dataFile = config.dataFile
outputFolder = config.localOutputFolderPath


def getInputFileList(inputFolder):
    fileList = {"id": "Filepath"}
    count = 0
    # File lookup in 0_Input Folder
    for root, directories, files in os.walk(inputFolder, topdown=False):
        for name in files:
            # print("Name in files | " + os.path.join(root, name))
            count = count + 1
            fileList[count] = os.path.join(root, name)
            # print(fileList.get(count))
            # sourceFilePath = fileList.get(count)

        """
        for name in directories:
            # print("Name in folders | " + os.path.join(root, name))
            fileList[count] = fileList[count].path.join(root, name)
            # print("Directory", fileList)
        """

    return fileList


def finalizedResultSet(resultSet, basename, mode):
    # mode - 1 = Single, 2 = Multiple
    if mode == 1:
        tmpList = resultSet
        del tmpList[0]

        result_dict = {"CustomerId": tmpList[0]["CustomerId"], "InvoiceDate": tmpList[0]["InvoiceDate"],
                       "InvoiceID": tmpList[0]["InvoiceID"], "InvoiceTotal": tmpList[0]["TotalAmount"],
                       "PurchaseOrder": tmpList[0]["PONum"],
                       "RemittanceAddressRecipient": tmpList[0]["BeneficiaryName"],
                       "VendorName": tmpList[0]["CompanyName"]}

        resultSet_list = [result_dict]

        # Audit message
        msg = 'File name : ' + basename + ' | Extraction result : ' + str(result_dict)
        insertAuditMsg('Data Extraction', msg)

        return resultSet_list, tmpList[0]["InvoiceID"]

    if mode == 2:

        FileList = getInputFileList(outputFolder)

        result_list = []
        tmpList = resultSet
        count = 0
        fileCount = count + 1
        for i in tmpList:
            # print(tmpList[count]["InvoiceID"])
            resultMsg = \
                {"CustomerId": tmpList[count]["CustomerId"], "InvoiceDate": tmpList[count]["InvoiceDate"],
                 "InvoiceID": tmpList[count]["InvoiceID"], "InvoiceTotal": tmpList[count]["TotalAmount"],
                 "PurchaseOrder": tmpList[count]["PONum"],
                 "RemittanceAddressRecipient": tmpList[count]["BeneficiaryName"],
                 "VendorName": tmpList[count]["CompanyName"]}

            result_list.append(resultMsg)

            # Audit message
            file_name = FileList.get(fileCount)
            basename = os.path.basename(file_name)
            msg = 'File name : ' + basename + ' | Extraction result : ' + str(resultMsg)
            insertAuditMsg('Data Extraction', msg)

            count = count + 1
            fileCount = fileCount + 1

        return result_list


def writeJsonOutput(lst, outputFolder, filename, mode):
    # mode - 1 = Single, 2 = Multiple
    # print(lst)
    if mode == 1:
        invoice_id = lst.get("InvoiceID")
        # print(invoice_id)
        file_name = str(filename) + '_' + str(invoice_id)
        outputFilePath = outputFolder + file_name
    if mode == 2:
        file_name = str(filename)
        outputFilePath = outputFolder + file_name

    jsonString = json.dumps(lst, indent=2)
    jsonFile = open(outputFilePath + ".json", "w")
    jsonFile.write(jsonString)
    jsonFile.close()

    return file_name + ".json"


def preProcessChecker(inputFolder):
    fileList = getInputFileList(inputFolder)  # file(s) lookup in Input Folder
    fileListLen = len(fileList)

    if (fileListLen - 1) == 0:
        print("NO Invoice File(s) Exist in the Input Folder")
        return False, fileListLen, fileList
    else:
        return True, fileListLen, fileList


def getSeqKey(table_name):
    book = dataFile
    b = openpyxl.load_workbook(book)
    table = b[table_name]

    i = 2
    while True:
        cellValue = table['A' + str(i)].value
        if cellValue is None:
            break
        else:
            i = i + 1

    b.close()
    return i


def insertRecord(result_list, mode):
    book = dataFile
    today = datetime.date.today()
    wb = openpyxl.load_workbook(book)
    table_name = 'Invoice'
    audit_table = 'Audit'

    seqKey = getSeqKey(table_name)
    auditKey = getSeqKey(audit_table)
    idx = seqKey - 1
    auditIdx = auditKey - 1

    if mode == 'single':
        table = wb[table_name]
        # print(seqKey, idx, a["VendorName"], a["PurchaseOrder"], a["InvoiceID"], a["InvoiceTotal"], today)
        table['A' + str(seqKey)] = idx
        table['B' + str(seqKey)] = result_list.get("VendorName")
        table['C' + str(seqKey)] = result_list.get("PurchaseOrder")
        table['D' + str(seqKey)] = result_list.get("InvoiceID")
        table['E' + str(seqKey)] = result_list.get("InvoiceTotal")
        table['F' + str(seqKey)] = today

        table = wb[audit_table]
        table['A' + str(auditKey)] = auditIdx
        table['B' + str(auditKey)] = today
        table['C' + str(auditKey)] = 'Record Insertion'
        msg = 'Index ' + str(idx) + ' | Vendor : [' + str(result_list.get("VendorName")) + '] | Invoice ID : [' + \
              str(result_list.get("InvoiceID")) + '] | PO Number : [' + str(result_list.get("PurchaseOrder")) + ']'
        table['D' + str(auditKey)] = msg

        idx = idx + 1
        seqKey = seqKey + 1
        auditKey = auditKey + 1
        auditIdx = auditIdx + 1

    else:
        for a in result_list:
            table = wb[table_name]
            # print(seqKey, idx, a["VendorName"], a["PurchaseOrder"], a["InvoiceID"], a["InvoiceTotal"], today)
            table['A' + str(seqKey)] = idx
            table['B' + str(seqKey)] = a["VendorName"]
            table['C' + str(seqKey)] = a["PurchaseOrder"]
            table['D' + str(seqKey)] = a["InvoiceID"]
            table['E' + str(seqKey)] = a["InvoiceTotal"]
            table['F' + str(seqKey)] = today

            table = wb[audit_table]
            table['A' + str(auditKey)] = auditIdx
            table['B' + str(auditKey)] = today
            table['C' + str(auditKey)] = 'Record Insertion'
            msg = 'Index' + str(idx) + ' | Vendor : [' + str(a["VendorName"]) + '] | Invoice ID : [' + \
                  str(a["InvoiceID"]) + '] | PO Number : [' + str(a["PurchaseOrder"] + ']')
            table['D' + str(auditKey)] = msg

            idx = idx + 1
            seqKey = seqKey + 1
            auditKey = auditKey + 1
            auditIdx = auditIdx + 1

    wb.save(book)
    wb.close()


def insertAuditMsg(category, msg):
    book = dataFile
    today = datetime.date.today()
    wb = openpyxl.load_workbook(book)
    audit_table = 'Audit'

    auditKey = getSeqKey(audit_table)
    auditIdx = auditKey - 1

    table = wb[audit_table]
    table['A' + str(auditKey)] = auditIdx
    table['B' + str(auditKey)] = today
    table['C' + str(auditKey)] = category
    table['D' + str(auditKey)] = msg

    wb.save(book)
    wb.close()


def validateExtractedRecord(newResult_list, mode):
    book = dataFile
    wb = openpyxl.load_workbook(book)
    invoice_table = 'Invoice'
    audit_table = 'Audit'
    today = datetime.date.today()
    seqKey = getSeqKey(invoice_table)
    maxKey = seqKey
    auditKey = getSeqKey(audit_table)
    auditIdx = auditKey - 1

    if mode == 'single':
        table = wb[invoice_table]
        newValueKey1 = newResult_list.get("VendorName") + '||' + newResult_list.get("InvoiceID")
        newValueKey2 = newResult_list.get("InvoiceID") + '||' + newResult_list.get("InvoiceTotal")

        for b in range(2, maxKey):
            key1 = str(table['B' + str(b)].value) + '||' + str(table['D' + str(b)].value)
            key2 = str(table['D' + str(b)].value) + '||' + str(table['E' + str(b)].value)
            if newValueKey1 == key1:
                msg = 'Duplicated invoice ID [' + str(table['D' + str(b)].value) + \
                      '] from the same company are detected at record index ' + str(b - 1)

                print(msg)
                # print('Duplicated invoice ID', str(table['D' + str(b)].value),
                # 'from the same company are detected at record index', b)

                table = wb[audit_table]
                table['A' + str(auditKey)] = auditIdx
                table['B' + str(auditKey)] = today
                table['C' + str(auditKey)] = 'Validation'
                table['D' + str(auditKey)] = msg
                auditKey = auditKey + 1
                auditIdx = auditIdx + 1

                # Mark 'Is_Duplicated' flag
                table = wb['Invoice']
                table['G' + str(b)] = 'Y'
                table = wb[audit_table]

                if newValueKey2 == key2:

                    table = wb[invoice_table]
                    msg = '[SAME] total invoice amount | Record Index [' + str(b - 1) + '] | Invoice ID [' + \
                          str(newResult_list.get("InvoiceID")) + '] | Extracted Total Amount : [' + \
                          str(newResult_list.get("InvoiceTotal")) + '] | ' + \
                          'Record found Total Amount : [' + str(table['E' + str(b)].value) + ']'

                    print(msg)
                    table = wb[audit_table]
                    table['A' + str(auditKey)] = auditIdx
                    table['B' + str(auditKey)] = today
                    table['C' + str(auditKey)] = 'Validation'
                    table['D' + str(auditKey)] = msg
                    auditKey = auditKey + 1
                    auditIdx = auditIdx + 1

                else:
                    table = wb[invoice_table]
                    msg = '[DIFFERENT] total invoice amount | Record Index [' + str(b - 1) + '] | Invoice ID [' + \
                          str(newResult_list.get("InvoiceID")) + '] | Extracted Total Amount : [' + \
                          str(newResult_list.get("InvoiceTotal")) + '] | ' + \
                          'Record found Total Amount : [' + str(table['E' + str(b)].value) + ']'

                    print(msg)
                    table = wb[audit_table]
                    table['A' + str(auditKey)] = auditIdx
                    table['B' + str(auditKey)] = today
                    table['C' + str(auditKey)] = 'Validation'
                    table['D' + str(auditKey)] = msg
                    auditKey = auditKey + 1
                    auditIdx = auditIdx + 1

    else:
        for a in newResult_list:
            table = wb[invoice_table]
            newValueKey1 = a["VendorName"] + '||' + a["InvoiceID"]
            newValueKey2 = a["InvoiceID"] + '||' + a["InvoiceTotal"]
            # print(newValueKey1, newValueKey2)

            for b in range(2, maxKey):
                key1 = str(table['B' + str(b)].value) + '||' + str(table['D' + str(b)].value)
                key2 = str(table['D' + str(b)].value) + '||' + str(table['E' + str(b)].value)
                if newValueKey1 == key1:
                    msg = 'Duplicated invoice ID [' + str(table['D' + str(b)].value) + \
                          '] from the same company are detected at record index ' + str(b - 1)

                    print(msg)
                    table = wb[audit_table]
                    table['A' + str(auditKey)] = auditIdx
                    table['B' + str(auditKey)] = today
                    table['C' + str(auditKey)] = 'Validation'
                    table['D' + str(auditKey)] = msg
                    auditKey = auditKey + 1
                    auditIdx = auditIdx + 1

                    # Mark 'Is_Duplicated' flag
                    table = wb['Invoice']
                    table['G' + str(b)] = 'Y'
                    table = wb[audit_table]

                    if newValueKey2 == key2:

                        table = wb[invoice_table]
                        msg = '[SAME] total invoice amount | Record Index [' + str(b - 1) + '] | Invoice ID [' + \
                              str(a["InvoiceID"]) + '] | Extracted Total Amount : [' + str(a["InvoiceTotal"]) + '] | ' + \
                              'Record found Total Amount : [' + str(table['E' + str(b)].value) + ']'

                        print(msg)
                        # print('Invoice total amount are same\n', a["InvoiceTotal"], str(table['E' + str(b)].value))
                        table = wb[audit_table]
                        table['A' + str(auditKey)] = auditIdx
                        table['B' + str(auditKey)] = today
                        table['C' + str(auditKey)] = 'Validation'
                        table['D' + str(auditKey)] = msg
                        auditKey = auditKey + 1
                        auditIdx = auditIdx + 1


                    else:
                        table = wb[invoice_table]
                        msg = '[DIFFERENT] total invoice amount | Record Index [' + str(b - 1) + '] | Invoice ID [' + \
                              str(a["InvoiceID"]) + '] | Extracted Total Amount : [' + str(a["InvoiceTotal"]) + '] | ' + \
                              'Record found Total Amount : [' + str(table['E' + str(b)].value) + ']'

                        print(msg)
                        # print('Invoice total amount are same\n', a["InvoiceTotal"], str(table['E' + str(b)].value))
                        table = wb[audit_table]
                        table['A' + str(auditKey)] = auditIdx
                        table['B' + str(auditKey)] = today
                        table['C' + str(auditKey)] = 'Validation'
                        table['D' + str(auditKey)] = msg
                        auditKey = auditKey + 1
                        auditIdx = auditIdx + 1

    wb.save(book)
    wb.close()
